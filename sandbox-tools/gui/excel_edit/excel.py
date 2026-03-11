#!/usr/bin/env python3
"""
excel_edit — read and write Excel (.xlsx) files using openpyxl.
Input:  {
  "op": "read|write|formula",
  "file": "/work/data.xlsx",
  "sheet": "Sheet1",
  "cell": "A1",      # for read/write
  "value": "hello",  # for write
  "range": "A1:C3"   # for read range
}
Output: {"result": ..., "exit_code": 0}
"""

import json
import os

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed", "exit_code": 1}))
    raise SystemExit

WORK_DIR = "/work"


def main():
    raw = os.environ.get("TOOL_INPUT", "{}")
    try:
        inp = json.loads(raw)
    except json.JSONDecodeError as e:
        out({"error": str(e), "exit_code": 1})
        return

    op = inp.get("op", "read")
    file_path = inp.get("file", "")
    sheet_name = inp.get("sheet", None)

    if not file_path:
        out({"error": "file field is required", "exit_code": 1})
        return

    # Safety: keep within WORK_DIR
    abs_path = os.path.realpath(os.path.join(WORK_DIR, file_path.lstrip("/")))
    if not abs_path.startswith(WORK_DIR):
        out({"error": "path traversal", "exit_code": 1})
        return

    try:
        if op == "read":
            wb = openpyxl.load_workbook(abs_path, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            cell_ref = inp.get("cell", "")
            range_ref = inp.get("range", "")

            if range_ref:
                rows = [[ws[cell].value for cell in row] for row in ws[range_ref]]
                out({"result": rows, "exit_code": 0})
            elif cell_ref:
                out({"result": ws[cell_ref].value, "exit_code": 0})
            else:
                # Return all data as 2D array
                data = [[cell.value for cell in row] for row in ws.iter_rows()]
                out({"result": data, "exit_code": 0})

        elif op == "write":
            wb = openpyxl.load_workbook(abs_path) if os.path.exists(abs_path) else openpyxl.Workbook()
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            cell_ref = inp.get("cell", "A1")
            value = inp.get("value")
            ws[cell_ref] = value
            wb.save(abs_path)
            out({"result": f"wrote {value!r} to {cell_ref}", "exit_code": 0})

        else:
            out({"error": f"unknown op: {op}", "exit_code": 1})

    except Exception as e:
        out({"error": str(e), "exit_code": 1})


def out(d: dict):
    print(json.dumps(d))


if __name__ == "__main__":
    main()
